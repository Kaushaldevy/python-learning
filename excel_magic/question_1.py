import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Calculation of total number of elements per supplier
products_per_supplier = {}
values_per_supplier = {}
low_inventory_count_products = {}

for product_row in range(2, product_list.max_row +1):
    supplier_name = product_list.cell(product_row, 4).value

    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name) + 1
    else:
        products_per_supplier[supplier_name] = 1
print(products_per_supplier)

# Calculation of total number inventory items per supplier
for product_row in range(2, product_list.max_row +1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    current_inventory_value = values_per_supplier.get(supplier_name)

    if supplier_name in values_per_supplier:
        values_per_supplier[supplier_name] = current_inventory_value + inventory * price
    else:
        values_per_supplier[supplier_name] = inventory * price
print(values_per_supplier)

# Identifying the low stock items
for product_row in range(2, product_list.max_row + 1):
    inventory = product_list.cell(product_row, 2).value
    if inventory < 10:
        low_inventory_count_products[product_list.cell(product_row, 1).value] = inventory

print(low_inventory_count_products)

# Write the total inventory price
for product_row in range(2, product_list.max_row +1):
    inventory_value_cell = product_list.cell(product_row, 5)
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    inventory_value_cell.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")
